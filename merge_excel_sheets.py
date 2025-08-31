!pip install pandas openpyxl

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from google.colab import files
import io

def merge_excel_sheets(file_path):
    all_data = []
    all_columns = set()
    
    xl_file = pd.ExcelFile(file_path)
    print(f"Found {len(xl_file.sheet_names)} sheets: {xl_file.sheet_names}")
    
    for sheet_name in xl_file.sheet_names:
        try:
            df = None
            for header_pos in [1, 0, 2]:
                try:
                    temp_df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_pos)
                    if not temp_df.empty:
                        temp_df.columns = temp_df.columns.astype(str)
                        col_sample = [str(col).strip() for col in temp_df.columns[:5]]
                        if any(col in ['Data1', 'Tree', 'City', 'Quant', 'Street', 'Gush', 'כמות', 'מין העץ'] for col in col_sample):
                            df = temp_df
                            print(f"Found headers at row {header_pos} for sheet '{sheet_name}'")
                            break
                except:
                    continue
            
            if df is None:
                df = pd.read_excel(file_path, sheet_name=sheet_name, header=1)
                df.columns = df.columns.astype(str)
            
            if not df.empty:
                new_cols = []
                col_count = {}
                
                for i, col in enumerate(df.columns):
                    col_str = str(col).strip()
                    if col_str == 'nan' or col_str.startswith('Unnamed:') or col_str == '':
                        col_str = f'Column_{i+1}'
                    
                    if col_str in col_count:
                        col_count[col_str] += 1
                        col_str = f'{col_str}_{col_count[col_str]}'
                    else:
                        col_count[col_str] = 0
                    
                    new_cols.append(col_str)
                
                df.columns = new_cols
                all_columns.update(df.columns.tolist())
                df['source_sheet'] = sheet_name
                all_data.append(df)
                print(f"Processing sheet '{sheet_name}': {len(df)} rows")
                print(f"  Columns: {df.columns.tolist()[:8]}...")
                
        except Exception as e:
            print(f"Error in sheet '{sheet_name}': {e}")
    
    final_columns = ['source_sheet'] + sorted([col for col in all_columns if col != 'source_sheet'])
    
    col_count = {}
    unique_final_columns = []
    for col in final_columns:
        if col in col_count:
            col_count[col] += 1
            unique_final_columns.append(f'{col}_duplicate_{col_count[col]}')
        else:
            col_count[col] = 0
            unique_final_columns.append(col)
    
    print(f"Total unique columns: {len(unique_final_columns)}")
    
    merged_df = pd.DataFrame()
    for df in all_data:
        temp_df = df.copy()
        for col in unique_final_columns:
            if col not in temp_df.columns:
                temp_df[col] = ''
        
        temp_df = temp_df.reindex(columns=unique_final_columns, fill_value='')
        merged_df = pd.concat([merged_df, temp_df], ignore_index=True)
    
    return merged_df, unique_final_columns

def improve_column_names(df, columns):
    improved_names = []
    name_count = {}
    
    # First pass - get basic improved names
    basic_names = []
    for col in columns:
        if col == 'source_sheet':
            basic_names.append('Source_Sheet')
            continue
        
        improved_name = col
        
        if col.lower() in ['data1', 'data1name']:
            improved_name = 'License_Number'
        elif col.lower() in ['tree']:
            improved_name = 'Tree_Type'
        elif col.lower() in ['quant']:
            improved_name = 'Tree_Quantity'
        elif col.lower() in ['city', 'cityname']:
            improved_name = 'City'
        elif col.lower() in ['street']:
            improved_name = 'Street'
        elif col.lower() in ['homenumber']:
            improved_name = 'House_Number'
        elif col.lower() in ['gush']:
            improved_name = 'Land_Block'
        elif col.lower() in ['helka']:
            improved_name = 'Land_Parcel'
        elif col.lower() in ['fromdate']:
            improved_name = 'Start_Date'
        elif col.lower() in ['todate']:
            improved_name = 'End_Date'
        elif col.lower() in ['price']:
            improved_name = 'Price'
        elif col.lower() in ['siba']:
            improved_name = 'Reason_Code'
        elif col.lower() in ['sibatext']:
            improved_name = 'Reason_Description'
        elif col.lower() in ['peula']:
            improved_name = 'Action_Type'
        elif col.lower() in ['rname']:
            improved_name = 'Applicant_Name'
        elif col.lower() in ['ezor']:
            improved_name = 'Area'
        elif 'date' in col.lower():
            improved_name = 'Date'
        elif 'name' in col.lower():
            improved_name = 'Name'
        elif col.startswith('Column_'):
            improved_name = f'Data_{col.split("_")[1]}'
        elif col == 'כמות':
            improved_name = 'Quantity'
        elif col == 'מין העץ':
            improved_name = 'Tree_Species'
        elif col == 'סיבה':
            improved_name = 'Reason'
        elif col == 'פעולה':
            improved_name = 'Action'
        elif col == 'מספר רישיון':
            improved_name = 'License_ID'
        elif col == 'ישוב':
            improved_name = 'Settlement'
        elif col == 'אזור':
            improved_name = 'Region'
        else:
            improved_name = col.replace(' ', '_').replace('/', '_')
        
        basic_names.append(improved_name)
    
    # Second pass - detect data types and handle duplicates
    for i, (col, basic_name) in enumerate(zip(columns, basic_names)):
        # Determine data type based on sample data
        sample_data = df[col].dropna().head(20)
        data_type = "Text"
        
        if len(sample_data) > 0:
            # Check if mostly numeric
            numeric_count = 0
            for value in sample_data:
                try:
                    float(value)
                    numeric_count += 1
                except:
                    pass
            
            if numeric_count / len(sample_data) > 0.7:  # If more than 70% are numeric
                data_type = "Number"
            elif sample_data.dtype.name.startswith('datetime'):
                data_type = "Date"
            else:
                data_type = "Text"
        
        # Handle duplicates by adding data type
        if basic_name in name_count:
            name_count[basic_name] += 1
            if data_type == "Number":
                final_name = f'{basic_name}_Num_{name_count[basic_name]}'
            elif data_type == "Date":
                final_name = f'{basic_name}_Date_{name_count[basic_name]}'
            else:
                final_name = f'{basic_name}_Text_{name_count[basic_name]}'
        else:
            # Check if this name will be duplicated later
            future_duplicates = sum(1 for future_name in basic_names[i+1:] if future_name == basic_name)
            if future_duplicates > 0:
                name_count[basic_name] = 0
                if data_type == "Number":
                    final_name = f'{basic_name}_Number'
                elif data_type == "Date":
                    final_name = f'{basic_name}_Date'
                else:
                    final_name = f'{basic_name}_Text'
            else:
                final_name = basic_name
                name_count[basic_name] = 0
        
        improved_names.append(final_name)
    
    print(f"\nColumn name improvements with data types:")
    for old, new in zip(columns, improved_names):
        if old != new:
            print(f"  {old} -> {new}")
    
    return improved_names

def save_formatted_excel(df, columns, output_path):
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='merged_data', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['merged_data']
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center")
        
        for col_num, column in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            worksheet.column_dimensions[cell.column_letter].width = 15

print("Upload the file:")
uploaded = files.upload()

file_name = list(uploaded.keys())[0]
print(f"\nProcessing file: {file_name}")

try:
    merged_data, columns = merge_excel_sheets(file_name)
    
    print(f"\nSummary:")
    print(f"Total rows: {len(merged_data)}")
    print(f"Total columns: {len(columns)}")
    
    source_counts = merged_data['source_sheet'].value_counts()
    print(f"\nBreakdown by sheets:")
    for source, count in source_counts.items():
        print(f"  {source}: {count} rows")
    
    output_filename = 'merged_report_2024.xlsx'
    
    print(f"\nImproving column names based on content...")
    improved_columns = improve_column_names(merged_data, columns)
    merged_data.columns = improved_columns
    
    save_formatted_excel(merged_data, improved_columns, output_filename)
    
    print(f"\nFile created successfully with improved column names!")
    
    print("Downloading the merged file:")
    files.download(output_filename)
    
    print(f"\nSample of merged data (first 5 rows):")
    print(merged_data.head())
    
except Exception as e:
    print(f"Error: {e}")

print("\nDone!")
